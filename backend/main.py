import os
import io
import re
import json
import base64
import fitz
import datetime
import requests
import threading
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional
import urllib.request
from pydantic import BaseModel
import psycopg2
import psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from groq import Groq
from dotenv import load_dotenv
from contextlib import asynccontextmanager

load_dotenv()

GROQ_API_KEYS = [k.strip() for k in os.getenv("GROQ_API_KEY", "").split(",") if k.strip()]
groq_clients = [Groq(api_key=k) for k in GROQ_API_KEYS]

def call_groq_api(**kwargs):
    if not groq_clients:
        raise Exception("Tidak ada GROQ_API_KEY yang dikonfigurasi.")
        
    last_exception = None
    for i, client in enumerate(groq_clients):
        try:
            return client.chat.completions.create(**kwargs)
        except Exception as e:
            last_exception = e
            error_str = str(e).lower()
            if any(err in error_str for err in ["429", "rate limit", "rate_limit", "413", "503", "capacity", "over capacity", "service unavailable"]):
                print(f"Fallback (Key {i+1} Limit/Capacity): Mencoba kunci berikutnya... Error: {e}")
                continue
            break
            
    raise last_exception

# Ambil URL Database PostgreSQL (Neon) dari environment variable
DATABASE_URL = os.getenv("DATABASE_URL")

def init_db():
    if not DATABASE_URL:
        print("WARNING: DATABASE_URL belum diatur!")
        return
    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()
    
    # Buat tabel dokumen jika belum ada
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS dokumen (
            id SERIAL PRIMARY KEY,
            nama_klien TEXT,
            jenis_dokumen TEXT,
            nomor_identitas TEXT,
            nilai_proyek TEXT,
            obligee TEXT,
            pekerjaan TEXT,
            masa_berlaku TEXT,
            teks_dokumen TEXT,
            created_at TIMESTAMP,
            updated_at TIMESTAMP
        )
    """)
    
    # Auto-migration: pastikan kolom-kolom register baru tersedia
    cursor.execute("ALTER TABLE dokumen ADD COLUMN IF NOT EXISTS kode_jenis TEXT;")
    cursor.execute("ALTER TABLE dokumen ADD COLUMN IF NOT EXISTS tgl_terbit TEXT;")
    cursor.execute("ALTER TABLE dokumen ADD COLUMN IF NOT EXISTS tgl_awal TEXT;")
    cursor.execute("ALTER TABLE dokumen ADD COLUMN IF NOT EXISTS tgl_akhir TEXT;")
    cursor.execute("ALTER TABLE dokumen ADD COLUMN IF NOT EXISTS durasi_hk TEXT;")
    cursor.execute("ALTER TABLE dokumen ADD COLUMN IF NOT EXISTS env TEXT DEFAULT 'production';")
    
    # Buat tabel audit_logs jika belum ada
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id SERIAL PRIMARY KEY,
            doc_id INTEGER,
            catatan TEXT,
            created_at TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def sync_to_google_sheets(payload):
    def _worker():
        webhook_url = os.getenv("GOOGLE_SHEETS_WEBHOOK_URL", "https://script.google.com/macros/s/AKfycbyqR2iO8lRtSNnJQWWzUXqHAqSpLYF5w2E5I10E-LPsViQpVUBymdEMRzjG_BnIRcqX8g/exec")
        if not webhook_url:
            return
        try:
            res = requests.post(webhook_url, json=payload, timeout=10)
            print("GOOGLE SHEETS SYNC STATUS:", res.status_code, res.text[:200])
        except Exception as e:
            print("GOOGLE SHEETS SYNC NOTICE:", e)

    threading.Thread(target=_worker, daemon=True).start()

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: Inisialisasi Database Tables
    init_db()
    yield
    # Shutdown

app = FastAPI(title="Insurance CRM API", lifespan=lifespan)

# Setup CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class DocumentUpdate(BaseModel):
    nama_klien: Optional[str] = ""
    jenis_dokumen: Optional[str] = ""
    nomor_identitas: Optional[str] = ""
    nilai_proyek: Optional[str] = ""
    obligee: Optional[str] = ""
    pekerjaan: Optional[str] = ""
    masa_berlaku: Optional[str] = ""
    teks_dokumen: Optional[str] = ""
    kode_jenis: Optional[str] = ""
    tgl_terbit: Optional[str] = ""
    tgl_awal: Optional[str] = ""
    tgl_akhir: Optional[str] = ""
    durasi_hk: Optional[str] = ""
    env: Optional[str] = "production"

def get_db_connection():
    if not DATABASE_URL:
        raise HTTPException(status_code=500, detail="Database URL belum dikonfigurasi")
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def encode_image_bytes(image_bytes):
    return base64.b64encode(image_bytes).decode('utf-8')

def parse_robust_json(content, fallback_data):
    if not content:
        return fallback_data
    # 1. Clean markdown & think tags
    clean = re.sub(r"<think>.*?</think>", "", content, flags=re.DOTALL)
    if "```json" in clean:
        clean = clean.split("```json")[1].split("```")[0]
    elif "```" in clean:
        clean = clean.split("```")[1].split("```")[0]
        
    start_idx = clean.find('{')
    end_idx = clean.rfind('}')
    
    if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
        json_candidate = clean[start_idx:end_idx+1]
        try:
            parsed = json.loads(json_candidate.strip())
            for k in fallback_data:
                if k in parsed and parsed[k]:
                    fallback_data[k] = parsed[k]
            return fallback_data
        except Exception:
            pass
            
    # 2. Regex field extractor if JSON was truncated or malformed
    keys = [
        "kode_jenis", "jenis_jaminan", "nomor_jaminan", "nilai_jaminan", 
        "principal", "obligee", "pekerjaan", "tgl_terbit", "tgl_awal", 
        "tgl_akhir", "durasi_hk", "masa_berlaku", "teks_asli"
    ]
    for k in keys:
        match = re.search(rf'"{k}"\s*:\s*"([^"\\]*(?:\\.[^"\\]*)*)"', clean, re.DOTALL)
        if match:
            fallback_data[k] = match.group(1).replace('\\"', '"')
            
    if not fallback_data.get("teks_asli") or fallback_data["teks_asli"] == "-":
        fallback_data["teks_asli"] = clean.strip()
        
    return fallback_data

MONTH_MAP = {
    'januari': '01', 'jan': '01',
    'februari': '02', 'feb': '02',
    'maret': '03', 'mar': '03',
    'april': '04', 'apr': '04',
    'mei': '05', 'may': '05',
    'juni': '06', 'jun': '06',
    'juli': '07', 'jul': '07',
    'agustus': '08', 'ags': '08', 'agu': '08', 'agt': '08',
    'september': '09', 'sep': '09',
    'oktober': '10', 'okt': '10',
    'november': '11', 'nopember': '11', 'nov': '11', 'nop': '11',
    'desember': '12', 'des': '12'
}

def parse_dates_from_text(text_to_search):
    if not text_to_search:
        return []
    month_names = "|".join(MONTH_MAP.keys())
    named_pattern = re.compile(rf'(\d{{1,2}})\s+({month_names})\s+(\d{{4}})', re.IGNORECASE)
    
    found_dates = []
    for match in named_pattern.finditer(text_to_search):
        day = match.group(1).zfill(2)
        month_str = match.group(2).lower()
        month = MONTH_MAP.get(month_str, "01")
        year = match.group(3)
        found_dates.append(f"{day}/{month}/{year}")
        
    if not found_dates:
        num_pattern = re.compile(r'(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})')
        for match in num_pattern.finditer(text_to_search):
            day = match.group(1).zfill(2)
            month = match.group(2).zfill(2)
            year = match.group(3)
            found_dates.append(f"{day}/{month}/{year}")
            
    return found_dates

def rapikan_teks(teks_mentah):
    fallback_data = {
        "kode_jenis": "PB", "jenis_jaminan": "-", "nomor_jaminan": "-", "nilai_jaminan": "-", 
        "principal": "-", "obligee": "-", "pekerjaan": "-", 
        "tgl_terbit": "-", "tgl_awal": "-", "tgl_akhir": "-", "durasi_hk": "-",
        "masa_berlaku": "-", "teks_asli": teks_mentah
    }
    if len(teks_mentah.strip()) < 20: return fallback_data
    
    if not groq_clients:
        fallback_data["teks_asli"] = "WARNING: GROQ_API_KEY belum dikonfigurasi di .env Server!\n\n" + teks_mentah
        return fallback_data
        
    base_prompt = (
        "Anda adalah asisten AI ahli administrasi asuransi Surety Bond & Bank Garansi Indonesia.\n"
        "Tugas Anda mengekstrak informasi dokumen menjadi format JSON standar register asuransi:\n\n"
        "ATURAN KODE JENIS BOND (PENTING):\n"
        "- 'PB' = Jaminan Pelaksanaan (Performance Bond)\n"
        "- 'MB' = Jaminan Pemeliharaan (Maintenance Bond)\n"
        "- 'APB' = Jaminan Uang Muka (Advance Payment Bond)\n"
        "- 'BB' = Jaminan Penawaran (Bid Bond / Tender)\n\n"
        "PANDUAN EKSTRAKSI FIELD:\n"
        "- nomor_jaminan: Nomor Sertifikat Polis / Nomor Jaminan / SPPBJ / Nomor Permohonan\n"
        "- kode_jenis: Kode singkatan ('PB', 'MB', 'APB', atau 'BB')\n"
        "- jenis_jaminan: Nama lengkap jenis jaminan (contoh: 'PB - Jaminan Pelaksanaan', 'MB - Jaminan Pemeliharaan')\n"
        "- principal: Nama Perusahaan/Badan Hukum Pemohon/Terjamin/Penyedia (contoh: CV. ..., PT. ...)\n"
        "- obligee: Nama Penerima Jaminan/Pemilik Proyek/Pemilik Pekerjaan/Pejabat Pembuat Komitmen (PPK)/Dinas/Kementerian/BUMN/Pemberi Kerja\n"
        "- nilai_jaminan: Nilai uang jaminan (sertakan 'Rp' dan nominal lengkap, contoh: 'Rp 14.945.040,00')\n"
        "- pekerjaan: Nama kegiatan / pengadaan / proyek pekerjaan secara lengkap\n"
        "- tgl_terbit: Tanggal surat/dokumen ditandatangani/diterbitkan (format: DD/MM/YYYY atau -)\n"
        "- tgl_awal: Tanggal mulai berlaku jaminan (format: DD/MM/YYYY atau -)\n"
        "- tgl_akhir: Tanggal berakhirnya jaminan (format: DD/MM/YYYY atau -)\n"
        "- durasi_hk: Angka jumlah hari kalender / hari kerja (contoh: 50, 120, 180)\n"
        "- masa_berlaku: Ringkasan rentang tanggal/jangka waktu (contoh: '50 hari kalender (23 Juni 2026 s/d 11 Agustus 2026)')\n"
        "- teks_asli: Rapikan teks OCR dengan jarak baris/enter baru untuk tiap poin penomoran atau pergantian data\n\n"
        "Format respons HANYA JSON murni:\n"
        "{\n"
        '  "kode_jenis": "PB/MB/APB/BB",\n'
        '  "jenis_jaminan": "...",\n'
        '  "nomor_jaminan": "...",\n'
        '  "nilai_jaminan": "...",\n'
        '  "principal": "...",\n'
        '  "obligee": "...",\n'
        '  "pekerjaan": "...",\n'
        '  "tgl_terbit": "...",\n'
        '  "tgl_awal": "...",\n'
        '  "tgl_akhir": "...",\n'
        '  "durasi_hk": "...",\n'
        '  "masa_berlaku": "...",\n'
        '  "teks_asli": "..."\n'
        "}\n"
        "Jangan gunakan tag <think>. Jika kolom benar-benar tidak ada di teks, isi '-'.\n\n"
        "Teks Dokumen:\n" + teks_mentah
    )
    
    result_data = dict(fallback_data)
    try:
        chat = call_groq_api(
            messages=[{"role": "user", "content": base_prompt}],
            model="openai/gpt-oss-120b",
            temperature=0.1,
            max_tokens=4000
        )
        result_data = parse_robust_json(chat.choices[0].message.content, result_data)
    except Exception as e:
        print("ERROR AI EXTRACTION ATTEMPT 1:", e)
        
    # Auto-Retry Loop jika ada kolom penting yang masih '-' (Maksimal 3 kali percobaan bertarget)
    critical_keys = ["principal", "obligee", "nilai_jaminan", "pekerjaan", "masa_berlaku", "jenis_jaminan", "durasi_hk", "tgl_awal", "tgl_akhir"]
    max_retries = 3
    for attempt in range(2, max_retries + 1):
        missing_keys = [k for k in critical_keys if not result_data.get(k) or result_data[k] == "-"]
        if not missing_keys:
            break
            
        print(f"AUTO-RETRY {attempt}/{max_retries} for missing keys: {missing_keys}")
        retry_prompt = (
            f"PERHATIAN: Kolom penting berikut belum ditemukan pada pembacaan awal: {', '.join(missing_keys)}.\n"
            "Mohon teliti kembali teks dokumen di bawah ini secara mendalam baris demi baris.\n"
            "Petunjuk Khusus:\n"
            "- Jika 'obligee' belum ada: cari bagian PENERIMA JAMINAN, PEMILIK PROYEK, PEJABAT PEMBUAT KOMITMEN (PPK), DINAS, INSTANSI, atau PERUSAHAAN PEMBERI KERJA.\n"
            "- Jika 'principal' belum ada: cari bagian PEMOHON, TERJAMIN, PENYEDIA JASA, NAMA PERUSAHAAN (PT/CV).\n"
            "- Jika 'nilai_jaminan' belum ada: cari simbol 'Rp', nilai jaminan %, atau nominal angka kontrak.\n"
            "- Jika 'pekerjaan' belum ada: cari nama paket pekerjaan, pengadaan barang/jasa, atau pembangunan.\n"
            "- Jika 'durasi_hk' atau tanggal belum ada: cari angka hari kalender/kerja (HK) dan rentang tanggal (s/d).\n\n"
            "Berikan respons JSON HANYA untuk kolom-kolom yang masih kosong tersebut:\n"
            "{\n" + ",\n".join([f'  "{k}": "..."' for k in missing_keys]) + "\n}\n"
            "Teks Dokumen:\n" + teks_mentah
        )
        try:
            chat_retry = call_groq_api(
                messages=[{"role": "user", "content": retry_prompt}],
                model="openai/gpt-oss-120b",
                temperature=0.2,
                max_tokens=2000
            )
            retry_parsed = parse_robust_json(chat_retry.choices[0].message.content, {})
            for k in missing_keys:
                if retry_parsed.get(k) and retry_parsed[k] != "-":
                    result_data[k] = retry_parsed[k]
        except Exception as err:
            print(f"ERROR ON AUTO-RETRY {attempt}:", err)
            
    # POST-PROCESSING: Normalisasi tanggal dan durasi deterministik
    # 1. Jika tgl_awal atau tgl_akhir kosong/'-', cari dari masa_berlaku atau teks mentah
    if not result_data.get("tgl_awal") or result_data["tgl_awal"] == "-" or not result_data.get("tgl_akhir") or result_data["tgl_akhir"] == "-":
        dates_in_masa = parse_dates_from_text(result_data.get("masa_berlaku", ""))
        if len(dates_in_masa) >= 2:
            if not result_data.get("tgl_awal") or result_data["tgl_awal"] == "-":
                result_data["tgl_awal"] = dates_in_masa[0]
            if not result_data.get("tgl_akhir") or result_data["tgl_akhir"] == "-":
                result_data["tgl_akhir"] = dates_in_masa[1]
        else:
            all_dates = parse_dates_from_text(teks_mentah)
            if len(all_dates) >= 2:
                if not result_data.get("tgl_awal") or result_data["tgl_awal"] == "-":
                    result_data["tgl_awal"] = all_dates[0]
                if not result_data.get("tgl_akhir") or result_data["tgl_akhir"] == "-":
                    result_data["tgl_akhir"] = all_dates[1]

    # 2. Jika tgl_terbit kosong/'-', cari tanggal SPPBJ / Surat / atau gunakan tgl_awal
    if not result_data.get("tgl_terbit") or result_data["tgl_terbit"] == "-":
        sppbj_match = re.search(r'(?:SPPBJ|Surat|Tanggal|Tgl)[^\n\r]*?(\d{1,2}\s+(?:Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Nopember|Desember)\s+\d{4})', teks_mentah, re.IGNORECASE)
        if sppbj_match:
            parsed = parse_dates_from_text(sppbj_match.group(1))
            if parsed:
                result_data["tgl_terbit"] = parsed[0]
        elif result_data.get("tgl_awal") and result_data["tgl_awal"] != "-":
            result_data["tgl_terbit"] = result_data["tgl_awal"]

    # 3. Jika durasi_hk kosong/'-', cari angka hari
    if not result_data.get("durasi_hk") or result_data["durasi_hk"] == "-":
        dur_match = re.search(r'(\d+)\s*(?:HK|Hari|hari kalender|hari kerja)', result_data.get("masa_berlaku", "") + " " + teks_mentah, re.IGNORECASE)
        if dur_match:
            result_data["durasi_hk"] = dur_match.group(1)

    # 4. Jika nomor_jaminan kosong/'-', cari nomor SPPBJ / Surat / Jaminan / Permohonan
    if not result_data.get("nomor_jaminan") or result_data["nomor_jaminan"] == "-":
        no_patterns = [
            r'(?:SPPBJ|Polis|Jaminan|Sertifikat|Kontrak|Permohonan)\s*(?:No\.?|Nomor)\s*[:\.]?\s*([A-Za-z0-9\.\/\-]{6,})',
            r'(?:No\.?|Nomor)\s*(?:SPPBJ|Polis|Jaminan|Sertifikat|Kontrak|Permohonan)\s*[:\.]?\s*([A-Za-z0-9\.\/\-]{6,})',
            r'SPPBJ\s*No\.?\s*[:\.]?\s*([A-Za-z0-9\.\/\-]{6,})'
        ]
        for pat in no_patterns:
            m = re.search(pat, teks_mentah, re.IGNORECASE)
            if m:
                val = m.group(1).strip().strip('.')
                if not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', val):
                    result_data["nomor_jaminan"] = val
                    break

    return result_data


def extract_from_image_vision(base64_image):
    """Mengekstrak teks dari gambar dengan Groq Vision lalu menstrukturkannya dengan GPT-120B"""
    fallback_data = {
        "kode_jenis": "PB", "jenis_jaminan": "-", "nomor_jaminan": "-", "nilai_jaminan": "-", 
        "principal": "-", "obligee": "-", "pekerjaan": "-", 
        "tgl_terbit": "-", "tgl_awal": "-", "tgl_akhir": "-", "durasi_hk": "-",
        "masa_berlaku": "-", "teks_asli": ""
    }
    if not groq_clients:
        return fallback_data
        
    try:
        # Step 1: Gunakan Qwen Vision sebagai mesin OCR murni
        ocr_prompt = "Lakukan OCR pada gambar dokumen Surety Bond ini. Transkripsikan semua teks yang terlihat pada gambar secara lengkap, jelas, dan akurat."
        chat = call_groq_api(
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": ocr_prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                    ]
                }
            ],
            model="qwen/qwen3.6-27b",
            temperature=0.1,
            max_tokens=3000
        )
        raw_content = chat.choices[0].message.content
        
        # Bersihkan tag <think> jika ada
        clean_text = re.sub(r"<think>.*?</think>", "", raw_content, flags=re.DOTALL).strip()
        if not clean_text:
            clean_text = raw_content.replace("<think>", "").replace("</think>", "").strip()
            
        if len(clean_text) < 15:
            fallback_data["teks_asli"] = "Tidak ada teks yang dapat dibaca dari gambar."
            return fallback_data
            
        # Step 2: Kirim teks OCR ke GPT-120B untuk ekstraksi JSON rapi
        return rapikan_teks(clean_text)
        
    except Exception as e:
        print("ERROR GROQ VISION:", e)
        fallback_data["teks_asli"] = "Pengekstrakan gagal: " + str(e)
        return fallback_data


@app.post("/api/extract")
async def extract_document(file: UploadFile = File(...)):
    contents = await file.read()
    
    try:
        if file.filename.lower().endswith('.pdf'):
            doc = fitz.open("pdf", contents)
            teks_digital = ""
            for page in doc:
                teks_digital += page.get_text().strip() + "\n"
                
            if len(teks_digital.strip()) > 50:
                # Digital PDF (bisa di-select teksnya) - mencakup seluruh halaman
                data_ekstrak = rapikan_teks(teks_digital)
                return {"status": "success", "data": data_ekstrak}
            else:
                # PDF Scan (berisi gambar scan) -> Render dan OCR tiap halaman (hingga 5 lembar)
                combined_ocr_text = ""
                total_pages = min(len(doc), 5)
                for page_idx in range(total_pages):
                    page = doc[page_idx]
                    pix = page.get_pixmap(dpi=150)
                    img_bytes = pix.tobytes("jpeg")
                    b64_image = encode_image_bytes(img_bytes)
                    
                    try:
                        ocr_prompt = "Lakukan OCR pada gambar dokumen Surety Bond ini. Transkripsikan semua teks yang terlihat pada gambar secara lengkap, jelas, dan akurat."
                        chat = call_groq_api(
                            messages=[
                                {
                                    "role": "user",
                                    "content": [
                                        {"type": "text", "text": ocr_prompt},
                                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_image}"}}
                                    ]
                                }
                            ],
                            model="qwen/qwen3.6-27b",
                            temperature=0.1,
                            max_tokens=3000
                        )
                        raw_content = chat.choices[0].message.content
                        clean_text = re.sub(r"<think>.*?</think>", "", raw_content, flags=re.DOTALL).strip()
                        if not clean_text:
                            clean_text = raw_content.replace("<think>", "").replace("</think>", "").strip()
                        if clean_text:
                            combined_ocr_text += f"\n--- Halaman {page_idx + 1} ---\n" + clean_text
                    except Exception as ocr_err:
                        print(f"Error OCR Halaman {page_idx + 1}:", ocr_err)
                
                if combined_ocr_text.strip():
                    data_ekstrak = rapikan_teks(combined_ocr_text)
                    return {"status": "success", "data": data_ekstrak}
                else:
                    return {"status": "success", "data": {
                        "kode_jenis": "PB", "jenis_jaminan": "-", "nomor_jaminan": "-", "nilai_jaminan": "-", 
                        "principal": "-", "obligee": "-", "pekerjaan": "-", 
                        "tgl_terbit": "-", "tgl_awal": "-", "tgl_akhir": "-", "durasi_hk": "-",
                        "masa_berlaku": "-", "teks_asli": "Tidak ada teks yang berhasil diekstrak."
                    }}
                
        elif file.filename.lower().endswith(('.jpg', '.jpeg', '.png')):
            # Langsung kirim ke Groq Vision
            b64_image = encode_image_bytes(contents)
            data_ekstrak = extract_from_image_vision(b64_image)
            return {"status": "success", "data": data_ekstrak}
        else:
            raise HTTPException(status_code=400, detail="Format tidak didukung. Harap upload PDF, JPG, atau PNG.")
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents")
def get_documents(env: Optional[str] = "production"):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if env == "testing":
        cursor.execute("SELECT * FROM dokumen WHERE env='testing' ORDER BY id DESC")
    else:
        cursor.execute("SELECT * FROM dokumen WHERE env='production' OR env IS NULL ORDER BY id DESC")
    docs = cursor.fetchall()
    conn.close()
    return [dict(ix) for ix in docs]

@app.post("/api/documents")
def save_document(doc: DocumentUpdate):
    waktu_sekarang = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    doc_env = doc.env or "production"
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO dokumen (
            nama_klien, jenis_dokumen, nomor_identitas, nilai_proyek, 
            obligee, pekerjaan, masa_berlaku, teks_dokumen, 
            kode_jenis, tgl_terbit, tgl_awal, tgl_akhir, durasi_hk,
            created_at, updated_at, env
        ) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        doc.nama_klien, doc.jenis_dokumen, doc.nomor_identitas, doc.nilai_proyek, 
        doc.obligee, doc.pekerjaan, doc.masa_berlaku, doc.teks_dokumen, 
        doc.kode_jenis, doc.tgl_terbit, doc.tgl_awal, doc.tgl_akhir, doc.durasi_hk,
        waktu_sekarang, waktu_sekarang, doc_env
    ))
    conn.commit()
    conn.close()

    # Trigger Live Sync to Google Sheets
    sheets_payload = {
        "action": "INSERT",
        "env": doc_env,
        "sheet_name": "TESTING" if (doc_env == "testing") else "REGISTER SURETY BOND",
        "nomor_identitas": doc.nomor_identitas or "-",
        "no_polis": doc.nomor_identitas or "-",
        "kode_jenis": doc.kode_jenis or "PB",
        "jenis_bond": doc.kode_jenis or "PB",
        "nama_klien": doc.nama_klien or "-",
        "principal": doc.nama_klien or "-",
        "obligee": doc.obligee or "-",
        "pekerjaan": doc.pekerjaan or "-",
        "nilai_proyek": doc.nilai_proyek or "-",
        "nilai_bond": doc.nilai_proyek or "-",
        "tgl_terbit": doc.tgl_terbit or doc.tgl_awal or "-",
        "tgl_awal": doc.tgl_awal or "-",
        "tgl_akhir": doc.tgl_akhir or "-",
        "durasi_hk": doc.durasi_hk or "-"
    }
    sync_to_google_sheets(sheets_payload)

    return {"status": "success"}

@app.delete("/api/documents/{doc_id}")
def delete_document(doc_id: int):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT * FROM dokumen WHERE id=%s", (doc_id,))
        doc = cursor.fetchone()
        
        cursor.execute("DELETE FROM audit_logs WHERE doc_id=%s", (doc_id,))
        cursor.execute("DELETE FROM dokumen WHERE id=%s", (doc_id,))
        conn.commit()
        conn.close()
        
        # Sinkronisasi HAPUS ke Google Sheets
        if doc:
            doc_env = doc["env"] if "env" in doc.keys() and doc["env"] else "production"
            delete_payload = {
                "action": "DELETE",
                "env": doc_env,
                "sheet_name": "TESTING" if (doc_env == "testing") else "REGISTER SURETY BOND",
                "nomor_identitas": doc["nomor_identitas"],
                "no_polis": doc["nomor_identitas"],
                "nama_klien": doc["nama_klien"],
                "principal": doc["nama_klien"],
                "nilai_proyek": doc["nilai_proyek"]
            }
            sync_to_google_sheets(delete_payload)
            
        return {"status": "success"}
    except Exception as e:
        print("ERROR DELETE:", e)
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/documents/{doc_id}")
def update_document(doc_id: int, doc: DocumentUpdate):
    waktu_sekarang = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # 1. Ambil data lama
    cursor.execute("SELECT * FROM dokumen WHERE id=%s", (doc_id,))
    old_data = cursor.fetchone()
    if not old_data:
        conn.close()
        raise HTTPException(status_code=404, detail="Dokumen tidak ditemukan")
        
    old_data_dict = dict(old_data)
    
    # 2. Bandingkan data untuk mencari perubahan
    perubahan = []
    field_map = {
        "nama_klien": ("Nama Klien", doc.nama_klien),
        "jenis_dokumen": ("Jenis Jaminan", doc.jenis_dokumen),
        "nomor_identitas": ("Nomor Jaminan", doc.nomor_identitas),
        "nilai_proyek": ("Nilai Jaminan", doc.nilai_proyek),
        "obligee": ("Obligee", doc.obligee),
        "pekerjaan": ("Pekerjaan", doc.pekerjaan),
        "masa_berlaku": ("Masa Berlaku", doc.masa_berlaku),
        "kode_jenis": ("Kode Jenis Bond", doc.kode_jenis),
        "tgl_terbit": ("Tanggal Terbit", doc.tgl_terbit),
        "tgl_awal": ("Tanggal Awal", doc.tgl_awal),
        "tgl_akhir": ("Tanggal Akhir", doc.tgl_akhir),
        "durasi_hk": ("Durasi HK", doc.durasi_hk)
    }
    
    for key, (label, new_val) in field_map.items():
        old_val = old_data_dict.get(key)
        if str(old_val).strip() != str(new_val).strip():
            perubahan.append(f"{label} diubah dari '{old_val}' menjadi '{new_val}'")
            
    # Track Teks Asli
    if str(old_data_dict.get("teks_dokumen")).strip() != str(doc.teks_dokumen).strip():
        perubahan.append("Teks Asli Dokumen telah diubah / diedit secara manual")
            
    # 3. Jika ada perubahan, catat ke audit_logs
    if perubahan:
        catatan_lengkap = " | ".join(perubahan)
        cursor.execute("INSERT INTO audit_logs (doc_id, catatan, created_at) VALUES (%s, %s, %s)", (doc_id, catatan_lengkap, waktu_sekarang))
    
    # 4. Update tabel dokumen
    cursor.execute("""
        UPDATE dokumen 
        SET nama_klien=%s, jenis_dokumen=%s, nomor_identitas=%s, nilai_proyek=%s, 
            obligee=%s, pekerjaan=%s, masa_berlaku=%s, teks_dokumen=%s, 
            kode_jenis=%s, tgl_terbit=%s, tgl_awal=%s, tgl_akhir=%s, durasi_hk=%s,
            updated_at=%s
        WHERE id=%s
    """, (
        doc.nama_klien, doc.jenis_dokumen, doc.nomor_identitas, doc.nilai_proyek, 
        doc.obligee, doc.pekerjaan, doc.masa_berlaku, doc.teks_dokumen, 
        doc.kode_jenis, doc.tgl_terbit, doc.tgl_awal, doc.tgl_akhir, doc.durasi_hk,
        waktu_sekarang, doc_id
    ))
    
    conn.commit()
    conn.close()

    # 5. Sinkronisasi UPDATE ke Google Sheets
    doc_env = doc.env or old_data_dict.get("env") or "production"
    sheets_update_payload = {
        "action": "UPDATE",
        "env": doc_env,
        "sheet_name": "TESTING" if (doc_env == "testing") else "REGISTER SURETY BOND",
        "old_nomor_identitas": old_data_dict.get("nomor_identitas"),
        "nomor_identitas": doc.nomor_identitas or "-",
        "no_polis": doc.nomor_identitas or "-",
        "kode_jenis": doc.kode_jenis or "PB",
        "jenis_bond": doc.kode_jenis or "PB",
        "nama_klien": doc.nama_klien or "-",
        "principal": doc.nama_klien or "-",
        "obligee": doc.obligee or "-",
        "pekerjaan": doc.pekerjaan or "-",
        "nilai_proyek": doc.nilai_proyek or "-",
        "nilai_bond": doc.nilai_proyek or "-",
        "tgl_terbit": doc.tgl_terbit or "-",
        "tgl_awal": doc.tgl_awal or "-",
        "tgl_akhir": doc.tgl_akhir or "-",
        "durasi_hk": doc.durasi_hk or "-"
    }
    sync_to_google_sheets(sheets_update_payload)

    return {"status": "success", "perubahan_dicatat": len(perubahan) > 0}

@app.get("/api/documents/{doc_id}/logs")
def get_audit_logs(doc_id: int):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM audit_logs WHERE doc_id=%s ORDER BY id DESC", (doc_id,))
    logs = cursor.fetchall()
    conn.close()
    
    # Convert datetime objects to string if psycopg2 returns datetime
    result = []
    for log in logs:
        log_dict = dict(log)
        if isinstance(log_dict['created_at'], datetime.datetime):
             log_dict['created_at'] = log_dict['created_at'].strftime("%Y-%m-%d %H:%M:%S")
        result.append(log_dict)
    return result

@app.get("/api/documents/export/excel")
def export_documents_excel():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute("SELECT * FROM dokumen ORDER BY id ASC")
    docs = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "REGISTER SURETY BOND"

    # Title Header
    ws.merge_cells("A1:J1")
    ws["A1"] = "REGISTER LAPORAN SURETY BOND & BANK GARANSI"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Table Headers
    headers = [
        "NO. POLIS", "JENIS BOND", "PRINCIPAL (TERJAMIN)", "OBLIGEE (PENERIMA)", 
        "NAMA PEKERJAAN / PROYEK", "NILAI BOND", "TGL TERBIT", "TGL AWAL", "TGL AKHIR", "JUMLAH HK"
    ]

    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    ws.row_dimensions[2].height = 28
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for r_idx, doc in enumerate(docs, 3):
        d = dict(doc)
        no_polis = d.get("nomor_identitas") or "-"
        jenis_bond = d.get("kode_jenis") or ("MB" if "pemeliharaan" in str(d.get("jenis_dokumen", "")).lower() else "PB" if "pelaksanaan" in str(d.get("jenis_dokumen", "")).lower() else "APB" if "uang muka" in str(d.get("jenis_dokumen", "")).lower() else "BB" if "penawaran" in str(d.get("jenis_dokumen", "")).lower() else "PB")
        principal = d.get("nama_klien") or "-"
        obligee = d.get("obligee") or "-"
        pekerjaan = d.get("pekerjaan") or "-"
        nilai_bond = d.get("nilai_proyek") or "-"
        tgl_terbit = d.get("tgl_terbit") or d.get("tgl_awal") or "-"
        tgl_awal = d.get("tgl_awal") or "-"
        tgl_akhir = d.get("tgl_akhir") or "-"
        durasi_hk = d.get("durasi_hk") or "-"

        row = [no_polis, jenis_bond, principal, obligee, pekerjaan, nilai_bond, tgl_terbit, tgl_awal, tgl_akhir, durasi_hk]
        ws.row_dimensions[r_idx].height = 24
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            if r_idx % 2 == 0:
                cell.fill = alt_fill
            if c_idx in [1, 2, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Register_Surety_Bond_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
